"""One-off: cross-check TYPE FOR RERA IDW column between live sheet and accounts team reference file."""
import datetime
import openpyxl
from upload_to_sheets import get_gspread_client, DEFAULT_CREDENTIALS, MASTER_SHEET_ID

REF_PATH = r"C:\Users\Win11-A\Downloads\Copy of DPL Bank Statements 2026-27 (3).xlsx"

TAB_MAP = {
    "YES BANK - 0264": "YES Master 0264",
    "YES BANK - 2477": "YES CR Free 2477",
    "YES BANK - 0490": "YES IDW 0490",
    "YES BANK - 2314": "YES AH 2314",
    "YES BANK - 0377": "YES Rera 0377",
    "YES BANK - 2457": "YES AH IDW 2457",
}


def norm_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    for fmt in ("%d-%b-%Y", "%d-%b-%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return str(v).strip()


def norm_amt(v):
    if v is None or v == "":
        return 0.0
    try:
        return round(float(str(v).replace(",", "").strip()), 2)
    except ValueError:
        return 0.0


def load_reference(ws):
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] == "SL#" and "TXN DATE" in row:
            header_row_idx = i
            break
    hdr = [c for c in ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)][0]
    hdr = list(hdr)
    i_date = hdr.index("TXN DATE")
    i_cr = hdr.index("CREDITS")
    i_db = hdr.index("DEBITS")
    i_type = hdr.index("TYPE FOR RERA IDW")

    lookup = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if i_date >= len(row) or row[i_date] is None:
            continue
        key = (norm_date(row[i_date]), norm_amt(row[i_cr]) if i_cr < len(row) else 0.0, norm_amt(row[i_db]) if i_db < len(row) else 0.0)
        lookup.setdefault(key, []).append(row[i_type] if i_type < len(row) else None)
    return lookup


def main():
    client = get_gspread_client(DEFAULT_CREDENTIALS)
    ss = client.open_by_key(MASTER_SHEET_ID)
    wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True)

    total_mismatches = 0
    for live_tab, ref_tab in TAB_MAP.items():
        ws_live = ss.worksheet(live_tab)
        ws_ref = wb[ref_tab]
        ref_lookup = load_reference(ws_ref)

        all_vals = ws_live.get_all_values()
        hdr = all_vals[0]
        i_date = hdr.index("TXN DATE")
        i_cr = hdr.index("CREDITS")
        i_db = hdr.index("DEBITS")
        i_type = hdr.index("TYPE FOR RERA IDW")
        i_desc = hdr.index("DESCRIPTION")

        mismatches = []
        for r_idx, row in enumerate(all_vals[1:], start=2):
            if len(row) <= i_date or not row[i_date]:
                continue
            key = (norm_date(row[i_date]), norm_amt(row[i_cr]), norm_amt(row[i_db]))
            candidates = ref_lookup.get(key)
            if not candidates:
                continue
            ref_type = candidates.pop(0)
            live_type = row[i_type] if len(row) > i_type else ""
            ref_type_norm = (ref_type or "").strip()
            live_type_norm = (live_type or "").strip()
            if ref_type_norm != live_type_norm:
                mismatches.append((r_idx, row[i_desc][:60], live_type_norm, ref_type_norm))

        print(f"\n=== {live_tab} ({len(mismatches)} mismatches) ===")
        for r_idx, desc, live_v, ref_v in mismatches:
            print(f"  row {r_idx}: '{desc}' | live='{live_v}' vs ref='{ref_v}'")
        total_mismatches += len(mismatches)

    print(f"\nTOTAL mismatches across all tabs: {total_mismatches}")


if __name__ == "__main__":
    main()
